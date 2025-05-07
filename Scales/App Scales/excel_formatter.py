import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import os
import tkinter as tk  # Убедитесь, что tkinter импортирован

def append_to_excel(filename, data_dict):
    try:
        if not os.path.exists(filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(list(data_dict.keys()))  # Заголовки столбцов
            ws.append(list(data_dict.values()))  # Данные
            wb.save(filename)
        else:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            ws.append(list(data_dict.values()))
            wb.save(filename)

        # Вызываем функцию форматирования после добавления данных
        format_excel(filename)

    except Exception as e:
        text_area.insert(tk.END, f"Ошибка при записи в файл {filename}: {e}")

def format_excel(filename):
    # Открываем существующий Excel файл
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Задаем цвет для колонок
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Желтый цвет
    fill_green = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # Зеленый цвет
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Красный цвет   
    fill_blue = PatternFill(start_color='00B7EB', end_color='00B7EB', fill_type='solid')  # Голубой цвет 

    # Устанавливаем границы
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Применяем цвет и границы к ячейкам
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.column in [1, 2]:   
                cell.fill = fill_yellow
            elif cell.column in [3, 4]:   
                cell.fill = fill_green
            elif cell.column in [5, 6]:  
                cell.fill = fill_red
            elif cell.column in [7, 8]: 
                cell.fill = fill_blue
            elif cell.column == 9:  
                cell.fill = fill_green
            elif cell.column == 10: 
                cell.fill = fill_red


            cell.border = thin_border  # Применяем границы ко всем ячейкам

    # Сохраняем файл с примененным форматированием
    workbook.save(filename)