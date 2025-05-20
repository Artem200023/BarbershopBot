import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import os
import tkinter as tk
from typing import Dict, Optional


class ExcelFormatter:
    """Класс для работы с Excel файлами: добавление данных и форматирование"""
    
    COLORS = {
        'yellow': 'FFFF00',
        'green': '008000',
        'red': 'FF0000',
        'blue': '00B7EB'
    }
    
    COLUMN_FORMATTING = {
        1: 'yellow', 2: 'yellow',    # Дата и Время
        3: 'green', 4: 'green',      # Замер норма и Вес норма
        5: 'red', 6: 'red',          # Замер не норма и Вес не норма
        7: 'blue', 8: 'blue',        # Кол-во замеров и Весь суммарный вес
        9: 'green',                   # Суммарный вес норма
        10: 'red'                     # Суммарный вес не норма
    }

    def __init__(self, filename: str, text_area: Optional[tk.Text] = None):
        self.filename = filename
        self.text_area = text_area
        self._setup_styles()

    def _setup_styles(self):
        """Инициализация стилей"""
        self.fills = {
            name: PatternFill(
                start_color=color, 
                end_color=color, 
                fill_type='solid'
            ) for name, color in self.COLORS.items()
        }
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def log_error(self, message: str):
        """Логирование ошибок в текстовое поле"""
        if self.text_area:
            self.text_area.insert(tk.END, f"Ошибка: {message}\n")

    def append_to_excel(self, data_dict: Dict):
        """Добавление данных в Excel файл"""
        try:
            if not os.path.exists(self.filename):
                self._create_new_file(data_dict)
            else:
                self._append_to_existing_file(data_dict)
            
            self.format_excel()

        except Exception as e:
            self.log_error(f"при записи в файл {self.filename}: {e}")

    def _create_new_file(self, data_dict: Dict):
        """Создание нового файла с данными"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(list(data_dict.keys()))  # Заголовки
        ws.append(list(data_dict.values()))  # Данные
        wb.save(self.filename)

    def _append_to_existing_file(self, data_dict: Dict):
        """Добавление данных в существующий файл"""
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        ws.append(list(data_dict.values()))
        wb.save(self.filename)

    def format_excel(self):
        """Форматирование Excel файла"""
        try:
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                     min_col=1, max_col=sheet.max_column):
                for cell in row:
                    self._apply_formatting(cell)

            workbook.save(self.filename)
        except Exception as e:
            self.log_error(f"при форматировании файла {self.filename}: {e}")

    def _apply_formatting(self, cell):
        """Применение форматирования к ячейке"""
        color_name = self.COLUMN_FORMATTING.get(cell.column)
        if color_name:
            cell.fill = self.fills[color_name]
        cell.border = self.thin_border

# Создаем глобальный экземпляр форматтера с None в качестве text_area
_default_formatter = ExcelFormatter('Datas.xlsx')

# Функции для обратной совместимости
def append_to_excel(filename, data_dict, text_area=None):
    formatter = ExcelFormatter(filename, text_area)
    formatter.append_to_excel(data_dict)

def format_excel(filename, text_area=None):
    formatter = ExcelFormatter(filename, text_area)
    formatter.format_excel()